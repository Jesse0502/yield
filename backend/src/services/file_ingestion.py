"""
File ingestion and RAG indexing service.

Handles:
- Uploading files to Supabase Storage
- Extracting text content
- Chunking with RecursiveCharacterTextSplitter
- Embedding chunks with the same Gemini embeddings used in tools.py
- Writing chunks into the `memories` table for retrieval.
"""

import uuid
import io
import csv
import logging
from pathlib import Path
from typing import Dict, Any, List

from fastapi import UploadFile, HTTPException

from langchain_text_splitters import RecursiveCharacterTextSplitter

from src.db.client import supabase
from src.services.tools import embeddings

logger = logging.getLogger("yield.services.file_ingestion")

# Optional imports for file type support
try:
    from pypdf import PdfReader
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


BUCKET_NAME = "memory-files"


async def ingest_file_to_memory(user_id: str, file: UploadFile) -> Dict[str, Any]:
    """
    Ingest a user-uploaded file into Supabase Storage and the memories vector store.

    Steps:
    - Upload raw file to Supabase Storage (bucket: memory-files)
    - Extract text (supports .txt, .md, .pdf, .csv, .xlsx, .xls)
    - Split into chunks via RecursiveCharacterTextSplitter
    - Embed each chunk with Gemini embeddings
    - Insert each chunk as a memory row in `memories` table with metadata
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Uploaded file has no filename")

    filename = file.filename
    ext = Path(filename).suffix.lower()

    # Read file bytes
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    logger.info("Ingesting file '%s' (%s) for user %s", filename, ext, user_id)

    # Upload to Supabase Storage
    file_id = str(uuid.uuid4())
    storage_path = f"{user_id}/{file_id}{ext}"

    try:
        # Supabase Storage expects raw bytes, not BytesIO
        supabase.storage.from_(BUCKET_NAME).upload(
            storage_path,
            file_bytes,
            file_options={"content-type": file.content_type or "application/octet-stream"},
        )
        logger.info("Uploaded file '%s' to storage path %s", filename, storage_path)
    except Exception as e:
        logger.exception("Failed to upload file '%s' to storage path %s", filename, storage_path)
        raise HTTPException(status_code=500, detail=f"Failed to upload file to storage (bucket '{BUCKET_NAME}'). Make sure the bucket exists and is public/private as expected. Original error: {e}")

    # Extract text based on file type
    text = None
    
    if ext in [".txt", ".md"]:
        # Plain text and markdown files
        try:
            text = file_bytes.decode("utf-8", errors="ignore")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to decode text file: {e}")
    
    elif ext == ".pdf":
        # PDF files
        if not PDF_SUPPORT:
            raise HTTPException(
                status_code=400,
                detail="PDF support requires 'pypdf' package. Please install it: pip install pypdf"
            )
        try:
            pdf_file = io.BytesIO(file_bytes)
            reader = PdfReader(pdf_file)
            text_parts = []
            for page in reader.pages:
                text_parts.append(page.extract_text())
            text = "\n\n".join(text_parts)
            if not text.strip():
                raise HTTPException(status_code=400, detail="PDF file appears to be empty or contains no extractable text")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to extract text from PDF: {e}")
    
    elif ext == ".csv":
        # CSV files
        try:
            csv_file = io.StringIO(file_bytes.decode("utf-8", errors="ignore"))
            reader = csv.reader(csv_file)
            rows = []
            for row in reader:
                rows.append(", ".join(row))
            text = "\n".join(rows)
            if not text.strip():
                raise HTTPException(status_code=400, detail="CSV file appears to be empty")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to extract text from CSV: {e}")
    
    elif ext in [".xlsx", ".xls"]:
        # Excel files
        if not EXCEL_SUPPORT:
            raise HTTPException(
                status_code=400,
                detail="Excel support requires 'openpyxl' package. Please install it: pip install openpyxl"
            )
        try:
            excel_file = io.BytesIO(file_bytes)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            text_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
                text_parts.append("")  # Empty line between sheets
            text = "\n".join(text_parts)
            if not text.strip():
                raise HTTPException(status_code=400, detail="Excel file appears to be empty")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to extract text from Excel file: {e}")
    
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: {ext}. Supported formats: .txt, .md, .pdf, .csv, .xlsx, .xls"
        )

    # Split into chunks for RAG
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=1000,
        chunk_overlap=200,
        separators=["\n\n", "\n", " ", ""],
    )
    chunks: List[str] = splitter.split_text(text)

    if not chunks:
        raise HTTPException(status_code=400, detail="No text content found to index in this file.")
    logger.info("Extracted %s chunks from file '%s'", len(chunks), filename)

    # Embed chunks using the same embeddings as memory tools
    try:
        vectors = await embeddings.aembed_documents(chunks)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to embed file chunks: {e}")

    # Prepare rows for insertion into memories table
    rows: List[Dict[str, Any]] = []
    for idx, (chunk_text, embedding) in enumerate(zip(chunks, vectors)):
        rows.append(
            {
                "content": chunk_text,
                "embedding": embedding,
                "user_id": user_id,
                "metadata": {
                    "source": "file",
                    "file_id": file_id,
                    "file_name": filename,
                    "storage_path": storage_path,
                    "chunk_index": idx,
                },
            }
        )

    # Insert all chunks in one go
    try:
        supabase.table("memories").insert(rows).execute()
    except Exception as e:
        logger.exception("Failed to insert %s chunks for file '%s'", len(rows), filename)
        raise HTTPException(status_code=500, detail=f"Failed to insert memory chunks: {e}")

    logger.info("Successfully ingested file '%s' with %s chunks", filename, len(chunks))
    return {
        "file_id": file_id,
        "file_name": filename,
        "storage_path": storage_path,
        "chunks_indexed": len(chunks),
    }


